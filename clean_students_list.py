
import os
import openpyxl

FILE_NAME = '2018_all_students.xlsx'
NEW_FILE_NAME = 'students_new.xlsx'

ALL_STUDENTS_SHEET = '總名單'
NEW_ALL_STUDENTS_SHEET = 'all_students'

STUDENT_EMAIL_COL_IN_ALL_STUDENTS_SHEET = 4

DISMISSED_STUDENTS_SHEETS = [
    '2018 開除名單',
    '2016開除名單',
    '2015新生開除名單'
]

STUDENT_EMAIL_COL_IN_DISMISSED_SHEETS = [
    6,
    2,
    2
]


def isDismissed(studentEmail, wb):
  for index in range(len(DISMISSED_STUDENTS_SHEETS)):
    sheet_name = DISMISSED_STUDENTS_SHEETS[index]
    sheet_dismissed = wb[sheet_name]
    row_count = sheet_dismissed.max_row
    for row in range(2, row_count + 1):
      if sheet_dismissed.cell(row=row, column=STUDENT_EMAIL_COL_IN_DISMISSED_SHEETS[index]).value == studentEmail:
        print('Student {} is dismissed in {}'.format(studentEmail, sheet_name))
        return True

  return False


def getRowCount(sheet, col):
  rowCount = 0
  max_row = sheet.max_row + 1

  for row in range(1, max_row):
    if sheet.cell(row=row, column=col).value is None and sheet.cell(row=row, column=col-1).value is None:
      break
    else:
      rowCount = rowCount + 1
  return rowCount


def statistics(wb):
  sheet_all = wb[ALL_STUDENTS_SHEET]
  # Print statistics information
  row_count_all = getRowCount(
      sheet_all, STUDENT_EMAIL_COL_IN_ALL_STUDENTS_SHEET)
  all_students_count = row_count_all - 1
  print("There are {} students in {}".format(
      all_students_count, ALL_STUDENTS_SHEET))

  dismissed_students_count = 0
  for index in range(len(DISMISSED_STUDENTS_SHEETS)):
    sheet_name = DISMISSED_STUDENTS_SHEETS[index]
    sheet_dismissed = wb[sheet_name]
    row_count = getRowCount(
        sheet_dismissed, STUDENT_EMAIL_COL_IN_DISMISSED_SHEETS[index])
    students_count = row_count - 1
    dismissed_students_count += students_count

    print("There are {} students in {}".format(students_count, sheet_name))

  print("There are {} students were dismissed".format(dismissed_students_count))


if __name__ == "__main__":
  wb = openpyxl.load_workbook(FILE_NAME)
  statistics(wb)

  sheet_all = wb[ALL_STUDENTS_SHEET]
  sheet_result = wb.create_sheet(title=NEW_ALL_STUDENTS_SHEET)

  # copy the field names in the first row
  column_count_all = sheet_all.max_column
  for col in range(1, column_count_all + 1):
    sheet_result.cell(row=1, column=col,
                      value=sheet_all.cell(row=1, column=col).value)

  row_result = 2
  # filter the students infor
  for row in range(2, sheet_all.max_row + 1):
    userEmail = sheet_all.cell(row=row, column=STUDENT_EMAIL_COL_IN_ALL_STUDENTS_SHEET).value
    if userEmail and isDismissed(userEmail, wb):
      continue
    for col in range(1, column_count_all + 1):
      sheet_result.cell(row=row_result, column=col,
                        value=sheet_all.cell(row=row, column=col).value)
    row_result += 1

  wb.save(NEW_FILE_NAME)
