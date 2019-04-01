import os
import openpyxl

FILE_NAME = '10k_locations.xlsx'
SHEET_NAME = 'locations'

COUNT = 10000

FIELDS = [
    'Admin1',
    'Admin2',
    'StreetAddress',
    'City',
    'Country',
    'PostalCode',
    'countryUnitCode',
    'countryScheme',
    'Latitude',
    'Longitude',
    'TIV',
    'TIVCurrency',
    'OccupancyScheme',
    'OccupancyType',
    'ConstructionScheme',
    'ConstructionClass',
    'NumOfStories',
    'YearBuilt',
    'Contract Excess',
    'Contract Excess Currency',
    'Contract Limit',
    'Contract Coverage',
    'Perils'
]

VALUES = [

]

if __name__ == "__main__":
  wb = openpyxl.Workbook()

  # sheet = wb.create_sheet(SHEET_NAME)
  sheet = wb.active

  # Create header
  for col in range(0, len(FIELDS)):
    sheet.cell(row=1, column=col+1, value=FIELDS[col])

  # Create locations
  for row in range(2, COUNT + 2):
    sheet.cell(row=row, column=1, value='CA')
    sheet.cell(row=row, column=2, value='Alameda')
    sheet.cell(row=row, column=3, value='7575 gateway blvd')
    sheet.cell(row=row, column=4, value='Newark')
    sheet.cell(row=row, column=7, value='USA')
    sheet.cell(row=row, column=8, value='94560')
    sheet.cell(row=row, column=9, value='')
    sheet.cell(row=row, column=10, value='')
    sheet.cell(row=row, column=11, value='')
    sheet.cell(row=row, column=12, value='')
    sheet.cell(row=row, column=13, value='')
    sheet.cell(row=row, column=14, value='USD')
    sheet.cell(row=row, column=15, value='AFM31')
    sheet.cell(row=row, column=16, value='ATC72A')
    sheet.cell(row=row, column=17, value='2')
    sheet.cell(row=row, column=18, value='1994')
    sheet.cell(row=row, column=19, value='12345')
    sheet.cell(row=row, column=20, value='USD')
    sheet.cell(row=row, column=21, value='2000')
    sheet.cell(row=row, column=22, value='20000')
    sheet.cell(row=row, column=23, value='EQ')

  wb.save(FILE_NAME)
