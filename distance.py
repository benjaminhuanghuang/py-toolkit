import os
import openpyxl
import requests
import json
from time import sleep
import time

from math import sin, cos, sqrt, atan2, radians, acos

FILE_NAME = '/Users/bhuang/OneDrive - RMS/Ben2018/dev-doc/Locations.xlsx'

SHEET_NAME = 'Locations'

COL_LONGITUDE = 2
COL_LATITUDE = 3

COL_DISTANCE1 = 6
COL_DISTANCE2 = 7

EARTH_RADIUS = 6371.0

'''
SELECT locationId, lat, long FROM Locations
WHERE (6371 * acos( cos( radians(-122) ) * cos( radians( lat ) ) *
cos( radians( 37 ) - radians(long) ) + sin( radians(-122) ) *
sin( radians(lat) ) )) <= 15000;
'''
def getDistance1(lng1, lat1, lng2, lat2):
  c = acos(cos(radians(lat1)) * cos(radians(lat2)) * cos(radians(lng1) - radians(lng2)) + sin(radians(lat1)) * sin(radians(lat2)))
  distance = EARTH_RADIUS * c

  return distance / 1.61

# https://www.movable-type.co.uk/scripts/latlong.html
def getDistance2(lng1, lat1, lng2, lat2):
  lat1_r = radians(lat1)
  lng1_r = radians(lng1)

  lat2_r = radians(lat2)
  lng2_r = radians(lng2)

  delt_lng_r = lng2_r - lng1_r
  delt_lat_r = lat2_r - lat1_r

  a = sin(delt_lat_r / 2)**2 + cos(lat1_r) * \
      cos(lat2_r) * sin(delt_lng_r / 2)**2
  c = 2 * atan2(sqrt(a), sqrt(1 - a))

  distance = EARTH_RADIUS * c
  distance = distance / 1.61
  return distance


def calcuateDistance():
  wb = openpyxl.load_workbook(FILE_NAME)
  sheet = wb[SHEET_NAME]

  lng_origin = sheet.cell(row=2, column=COL_LONGITUDE).value
  lat_origin = sheet.cell(row=2, column=COL_LATITUDE).value

  for row in range(3, sheet.max_row + 1):
    lng = sheet.cell(row=row, column=COL_LONGITUDE).value
    lat = sheet.cell(row=row, column=COL_LATITUDE).value
    name = sheet.cell(row=row, column=4).value

    print('Calculate distance between RMS and ' + name)
    distance1 = getDistance1(lng_origin, lat_origin, lng, lat)
    distance2 = getDistance2(lng_origin, lat_origin, lng, lat)

    sheet.cell(row=row, column=COL_DISTANCE1, value=distance1)
    sheet.cell(row=row, column=COL_DISTANCE2, value=distance2)

    row += 1

  wb.save(FILE_NAME)


def compareAlgorithem():
  # RMS
  lat1 = 37.5403123
  lng1 = -122.0640977
  # 
  lat2 = 38.5395407
  lng2 = -121.4868429
  # lat2 = 36.157735
  # lng2 = -115.201738

  start = time.time()
  for i in range(10000):
    getDistance1(lng1, lat1, lng2, lat2)
  end = time.time()
  print("Time used by method1: ", end - start)
  
  start = time.time()
  for i in range(10000):
    getDistance2(lng1, lat1, lng2, lat2)
  end = time.time()
  print("Time used by method2: ", end - start)
  

if __name__ == "__main__":
  # calcuateDistance()
  compareAlgorithem()
