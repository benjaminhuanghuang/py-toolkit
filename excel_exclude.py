
import os
import openpyxl


def isDismissed(studentNum, sheet):
  row_count = sheet.max_row
  for row in range(2, row_count + 1):
    if sheet_all.cell(row=row, column=4).value == studentNum:
      print('Student {} is dismissed'.format(studentNum))
      return True

  return False


if __name__ == "__main__":
  wb = openpyxl.load_workbook('test.xlsx')
  sheet_all = wb['all']
  sheet_dismissed = wb['kaichu']
  sheet_result = wb.create_sheet(title='result')

  row_count_all = sheet_all.max_row
  column_count_all = sheet_all.max_column
  print ("There are {} students.".format(row_count_all - 1))

  row_count_dismissed = sheet_dismissed.max_row
  print ("There are {} students were dismissed.".format(row_count_dismissed - 1))

  # copy the field names in first row
  for col in range(1, column_count_all + 1):
    sheet_result.cell(row=1, column=col,
                      value=sheet_all.cell(row=1, column=col).value)

  row_result = 2
  # filter the students infor
  for row in range(2, row_count_all + 1):
    if isDismissed(sheet_all.cell(row=row, column=4).value, sheet_dismissed):
      continue
    for col in range(1, column_count_all + 1):
      sheet_result.cell(row=row_result, column=col,
                        value=sheet_all.cell(row=row, column=col).value)
    row_result += 1

  wb.save('test_new.xlsx')
