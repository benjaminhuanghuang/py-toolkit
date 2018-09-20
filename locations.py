import os
import openpyxl
import requests
import json
from time import sleep

FILE_NAME = '/Users/bhuang/OneDrive - RMS/Ben2018/dev-doc/Locations.xlsx'

SHEET_NAME = 'Locations'

COL_ADDRESS = 1
COL_LNG = 2
COL_LAT = 3

# The key belongs to RMS developer
GOOGLE_API_KEY = 'AIzaSyDavRIt8TjnghoIjx8R2SzhFG6AKCj6WGk'
GOOGLE_API_URL = 'https://maps.googleapis.com/maps/api/geocode/json?address={}&key=AIzaSyDavRIt8TjnghoIjx8R2SzhFG6AKCj6WGk'


def getCoordinateOfAddress(address):
  print("Get locaiton of " + address)
  url = GOOGLE_API_URL.format(address)
  response = requests.get(url)
  data_json = json.loads(response.text)
  location = data_json['results'][0]['geometry']['location']

  return location['lng'], location['lat']


if __name__ == "__main__":
  wb = openpyxl.load_workbook(FILE_NAME)
  sheet = wb[SHEET_NAME]

  for row in range(2, sheet.max_row + 1):
    address = sheet.cell(row=row, column=COL_ADDRESS).value
    lng, lat = getCoordinateOfAddress(address)
    sheet.cell(row=row, column=COL_LNG, value=lng)
    sheet.cell(row=row, column=COL_LAT, value=lat)

    row += 1
    sleep(3)

  wb.save(FILE_NAME)
