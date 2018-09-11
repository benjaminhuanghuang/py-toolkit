import os
import openpyxl
# Import smtplib for the actual sending function
import smtplib

# Import the email modules we'll need
from email.mime.text import MIMEText

FILE_NAME = '2018_all_students.xlsx'


COL_NAME = 2
COL_EMAIL = 3

PATH_EMAIL_TEMPLATE = ''


if __name__ == "__main__":
  wb = openpyxl.load_workbook(FILE_NAME)
  sheet_students = wb['students']
  
  row_count = sheet_students.max_row

  # Send the message via SMTP server.
  s = smtplib.SMTP('localhost')
  
  with open(PATH_EMAIL_TEMPLATE) as fp:
    # Create a text/plain message
    text = fp.read()
    msg = MIMEText(text)
    for row in range(2, row_count + 1):
      name = sheet_students.cell(row=row, column=COL_NAME).value
      email_address = sheet_students.cell(row=row, column=COL_EMAIL).value
      msg['Subject'] = "Subject_Hello"
      msg['From'] = 'pirateliu'
      msg['To'] = email_address
      s.send_message(msg)
  s.quit()

    