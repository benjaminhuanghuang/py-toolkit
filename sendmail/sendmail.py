import os
import configparser # python 3
import openpyxl
# Import smtplib for the actual sending function
import smtplib, ssl

from email import encoders
# will contain the HTML and plain-text
from email.mime.text import MIMEText
# instance combines these into a single message with two alternative rendering options:
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase

# the excel sheet contains receivers 
EXCEL_NAME = 'receivers.xlsx'
SHEET_NAME = 'receivers'

# template 
TEXT_TEMPLATE = 'template.txt'
HTML_TEMPLATE = 'template.html'

# attachment file
ATTACHMENT = '「香柏木培訓中心」校規_2020.pdf'

# Should enable: Less secure app access
EMAIL_SERVER = 'smtp.gmail.com'

COL_NAME = 1
COL_EMAIL = 2
COL_SCORE = 5

def replaceTemplate(text, name, score):
  return text.replace("#name#", name).replace("#score#", str(score))


if __name__ == "__main__":
  # Read template
  htmlFile = open(HTML_TEMPLATE, "r")
  htmlContent = htmlFile.read()
  htmlFile.close()

  textFile = open(TEXT_TEMPLATE, "r")
  textContent = textFile.read()
  textFile.close()

  # Read receivers
  wb = openpyxl.load_workbook(EXCEL_NAME)
  sheet_receivers = wb[SHEET_NAME]

  row_count = sheet_receivers.max_row

  config = configparser.ConfigParser()
  config.read('config.ini')

  emailAccount = config['DEFAULT']['USER']
  password = config['DEFAULT']['PASSWORD']
  # print(username, password)
  
  # Attachment
  attachmentPart = MIMEBase("application", "octet-stream")
   # Open PDF file in binary mode
  with open(ATTACHMENT, "rb") as attachment:
      # Add file as application/octet-stream
      # Email client can usually download this automatically as attachment
      attachmentPart = MIMEBase("application", "octet-stream")
      attachmentPart.set_payload(attachment.read())

  # Encode file in ASCII characters to send by email    
  encoders.encode_base64(attachmentPart)

  # Add header as key/value pair to attachment part
  attachmentPart.add_header('Content-Disposition','attachment', filename=ATTACHMENT)
  

  # Create a secure SSL context
  context = ssl.create_default_context()
  # 465 (SSL required) or 587 (TLS required)
  with smtplib.SMTP("smtp.gmail.com", 587) as server: 
    server.ehlo()
    server.starttls()
    # Login SMTP server.
    # Have to enable https://myaccount.google.com/lesssecureapps?pli=1
    server.login(emailAccount, password)
    print("Loged in")
    attachment = open(ATTACHMENT, "rb")

    # Send the message via SMTP server.
    for row in range(1, row_count + 1):
      # Get receiver infor
      name = sheet_receivers.cell(row=row, column=COL_NAME).value
      email_address = sheet_receivers.cell(row=row, column=COL_EMAIL).value
      # score = sheet_receivers.cell(row=row, column=COL_SCORE).value
      print("Send email to {} {}".format(name, email_address))
      if(email_address):
        mainMessage = MIMEMultipart('mixed')
        mainMessage["Subject"] = config['DEFAULT']['SUBJECT']
        mainMessage["From"] = "{} <{}>".format(config['DEFAULT']['SENDER'],config['DEFAULT']['SENDER_EMAIL'])
        mainMessage["To"] = email_address

        message = MIMEMultipart('alternative')
     
        # Replace emial body
        textMessage = textContent.replace("#name#", name)
        textPart = MIMEText(textMessage, "plain")

        htmlMessage = htmlContent.replace("#name#", name)
        htmlPart = MIMEText(htmlMessage, "html")

        message.attach(textPart)
        message.attach(htmlPart)

        mainMessage.attach(message)
        mainMessage.attach(attachmentPart)

        server.sendmail(config['DEFAULT']['SENDER_EMAIL'], email_address, mainMessage.as_string())
      
  attachment.close()

